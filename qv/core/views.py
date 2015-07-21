from django.http import HttpResponseRedirect
from django.shortcuts import render, get_object_or_404
from core.forms import *
from core.models import Dataset
from core.utils import *
import csv
from operator import itemgetter
import os
from openpyxl import load_workbook
import json

def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            instance = Dataset(
                file_field=request.FILES['file']
                ,name=request.POST['dataset_name'])
            instance.save()
            #Success
    else:
        form = UploadFileForm()
    datasets = Dataset.objects.all()
    for dataset in datasets:
        ext = os.path.splitext(dataset.file_field.name)[1]
        if ext==".csv":
            dataset.data = ", ".join([row for row in csv.reader(dataset.file_field.read().splitlines())][0])
        elif (ext==".xls" or ext==".xlsx"):
            wb = load_workbook(filename=dataset.file_field)
            sheets = wb.get_sheet_names()
            dataset.data = ", ".join(sheets)
        else:
            dataset.data = "More config required"
    return render(request, 'upload.html', {'form': form,'datasets':datasets})

def data(request,slug):
    dataset = get_object_or_404(Dataset,slug=slug)
    ext = os.path.splitext(dataset.file_field.name)[1]
    oldSheet = dataset.sheet
    oldRow = dataset.starting_row
    if request.method == 'POST':
        form = ConfigDatasetForm(request.POST,instance=dataset)
        if form.is_valid():
            if (form.cleaned_data['sheet']!=oldSheet or form.cleaned_data['starting_row']!=oldRow) and (ext==".xls" or ext==".xlsx"):
                dataset.json = None
            form.save()
            #Success
    else:
        form = ConfigDatasetForm(instance=dataset)
    if dataset.json is None:
        if ext==".csv":
            dataArr = [row for row in csv.reader(dataset.file_field.read().splitlines())]
        elif ext==".xls" or ext==".xlsx":
            wb = load_workbook(filename=dataset.file_field)
            sheets = wb.get_sheet_names()
            if dataset.sheet is not None:
                ws = wb[sheets[dataset.sheet]]
            else:
                #Default is sheet 0
                ws = wb[sheets[0]]
            if dataset.starting_row is not None:
                dataArr = [[notNone(str(cell.value).encode('ascii',errors='ignore')) for cell in row] for row in ws.rows[dataset.starting_row:]]
            else:
                #Default is start from row 0
                dataArr = [[notNone(str(cell.value).encode('ascii',errors='ignore')) for cell in row] for row in ws.rows]
        else:
            #Something is wrong
            dataset.data = "More config required"
            data = {
                'dataset':dataset,
                'form':form
            }
            return render(request,'data.html',data)
        headers = dataArr[0]
        dataDict = []
        for row in dataArr[1:]:
            rowDict = {}
            for i in range(len(headers)):
                header = headers[i]
                cell = num(row[i])
                rowDict[header] = cell
            dataDict.append(rowDict)
        dataset.json = dataDict
        dataset.save()
        dataset.data = dataDict
        data = {
            'dataset':dataset,
            'form':form
        }
    else:
        dataset.data = byteify(dataset.json)
        data = {
            'dataset':dataset,
            'form':form
        }
    return render(request,'data.html',data)